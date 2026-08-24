"""
generic_tabular_parser.py — A format-agnostic fallback for tradebook/ledger
exports that don't match tradebook_parser.py's Zerodha-specific template.

Rather than hardcoding a parser per broker (which needs a real sample file
per broker to build and verify against — see tradebook_parser.py's own
DECISION notes on why that's the approach for Zerodha specifically), this
scans for a header row using fuzzy matching against known synonyms for the
columns any trade record needs: symbol, ISIN, date, quantity, price, and
buy/sell direction. If it finds a plausible header, it extracts rows below
it the same way regardless of the exact column wording or order a given
broker uses.

This sits BETWEEN the Zerodha-specific parser and the LLM fallback in
upload_dispatch.py's chain:
  1. tradebook_parser (Zerodha-exact) — fastest, most precise, zero cost
  2. generic_tabular_parser (this file) — heuristic, no API key needed,
     handles "reasonably standard" tradebook/ledger layouts from other
     brokers (Motilal Oswal, Angel One, etc.) without ever leaving the
     machine
  3. llm_parser (opt-in, needs an API key) — last resort for genuinely
     unusual layouts

DECISION: matching is a two-pass process, verified against two real broker
exports (not just synthetic test files) that turned out to use abbreviated,
prefixed column names neither exact-match nor naive substring-match handled
correctly:
  Pass 1 — exact match against the synonym list (highest confidence).
  Pass 2 — for any still-unmatched required field, a guarded fuzzy pass:
    the normalized header must START WITH a synonym (not just contain it
    anywhere), e.g. "scriptname" starts with "script" and "transtype"
    matches via a dedicated synonym rather than a risky generic prefix.
    The "symbol" field additionally excludes any header containing "code"
    or "id", since real exports frequently have a numeric internal
    "ScriptCode"/"StockCode" column alongside the actual name column, and a
    naive prefix match would grab the numeric code instead of the name.

DECISION: some brokers (confirmed against a real export) split price into
two columns — "Buy Price" and "Sell Price" — populating only the one that
applies to that row's direction, rather than one unified "Price"/"Rate"
column. This parser detects that pattern and treats it as satisfying the
price requirement, picking whichever of the two is actually populated on
each row.

DECISION: acceptance requires matching symbol + quantity + (price OR
buy/sell-price pair), AND at least one of (date, trade_type) — a
deliberately conservative bar, so a file that only coincidentally has a
column called "Price" doesn't get mis-parsed as a tradebook. If the bar
isn't met, this parser reports "no confident match" rather than guessing,
and the caller falls through to the next tier instead of trusting a
low-confidence extraction. The header row itself may appear many rows into
the sheet (a real export had 33 rows of charges/summary preamble first) —
every row is scanned, not just the first few.
"""
from __future__ import annotations

import io
import re
import uuid
import datetime as dt
from typing import List, Optional

import openpyxl

# --------------------------------------------------------------------------- #
# Column synonym groups — case-insensitive, punctuation/whitespace-normalized
# --------------------------------------------------------------------------- #

SYNONYMS = {
    "symbol": ["symbol", "scrip", "scripname", "scriptname", "stock", "stockname", "security",
               "securityname", "instrument", "tradingsymbol", "company", "companyname",
               "script", "scripcontract", "contract"],
    "isin": ["isin", "isincode", "isinno", "isinnumber"],
    "date": ["tradedate", "trandate", "date", "transactiondate", "orderdate", "txndate",
             "dealdate", "settlementdate"],
    "quantity": ["quantity", "qty", "shares", "units", "noofshares", "nos", "tradedqty", "delivqty"],
    "price": ["price", "rate", "tradeprice", "avgprice", "averageprice", "dealprice",
              "priceperunit", "netrate", "transrate"],
    "trade_type": ["tradetype", "type", "buysell", "transactiontype", "txntype", "transtype",
                    "bs", "action", "buyorsell", "orderindicator"],
    "buy_price": ["buyprice", "buyrate"],
    "sell_price": ["sellprice", "sellrate"],
}

# Fields where a prefix-fuzzy match (pass 2) is allowed to help — deliberately
# excludes "isin" (too short/risky to prefix-match) and treats "symbol"
# specially (see _fuzzy_match_field's code/id exclusion).
FUZZY_ELIGIBLE_FIELDS = {"symbol", "date", "quantity", "price", "trade_type", "buy_price", "sell_price"}

BUY_TOKENS = {"b", "buy", "bought", "purchase"}
SELL_TOKENS = {"s", "sell", "sold", "sale"}

_ISIN_RE = re.compile(r"^IN[A-Z0-9]{10}$")


def _normalize_header(text) -> str:
    if text is None:
        return ""
    return re.sub(r"[^a-z0-9]", "", str(text).lower())


def _fuzzy_match_field(norm_header: str, field: str) -> bool:
    """Pass-2 matcher: normalized header must START WITH a known synonym.
    'symbol' additionally rejects headers containing 'code' or 'id', since
    those are almost always a numeric internal identifier column sitting
    alongside the real name column, not the name itself."""
    if field == "symbol" and ("code" in norm_header or norm_header.endswith("id")):
        return False
    return any(norm_header.startswith(syn) for syn in SYNONYMS[field])


def _find_header_row(rows: List[tuple]) -> Optional[dict]:
    """Scans every row for the best-matching header row (a real export had
    33 rows of unrelated summary content before the actual header). Returns
    a dict with 'row_index' and 'columns' (field -> column-index) if a
    confident match is found, else None. Exact-match synonyms are tried
    first per column; unmatched required fields fall back to the guarded
    prefix-fuzzy pass."""
    best_map, best_score, best_idx = None, 0, -1
    for row_idx, row in enumerate(rows):
        col_map = {}
        # Pass 1: exact match
        for idx, cell in enumerate(row):
            norm = _normalize_header(cell)
            if not norm:
                continue
            for field, synonyms in SYNONYMS.items():
                if field in col_map:
                    continue
                if norm in synonyms:
                    col_map[field] = idx
                    break
        # Pass 2: guarded fuzzy prefix match, only for fields still missing
        for idx, cell in enumerate(row):
            norm = _normalize_header(cell)
            if not norm:
                continue
            for field in FUZZY_ELIGIBLE_FIELDS:
                if field in col_map:
                    continue
                if _fuzzy_match_field(norm, field):
                    col_map[field] = idx

        has_price = "price" in col_map or ("buy_price" in col_map and "sell_price" in col_map)
        required_core = "symbol" in col_map and "quantity" in col_map and has_price
        has_context = "date" in col_map or "trade_type" in col_map
        score = len(col_map)
        if required_core and has_context and score > best_score:
            best_score, best_map, best_idx = score, col_map, row_idx

    if best_map is None:
        return None
    return {"row_index": best_idx, "columns": best_map}
    return best_map


def _parse_date(value) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, (dt.date, dt.datetime)):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%d-%b-%Y", "%d %b %Y", "%m/%d/%Y"):
        try:
            return dt.datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s if s else None  # keep the raw string rather than silently dropping the row


def _parse_trade_type(value) -> Optional[str]:
    if value is None:
        return None
    token = str(value).strip().lower()
    if token in BUY_TOKENS:
        return "buy"
    if token in SELL_TOKENS:
        return "sell"
    return None


def _to_float(value) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = re.sub(r"[,\s₹]", "", str(value))
    try:
        return float(s)
    except ValueError:
        return None


def parse_tabular_bytes(filename: str, file_bytes: bytes) -> dict:
    """Returns {"transactions": [...], "warnings": [...], "error": str|None,
    "columns_matched": [...]}. Only supports .xlsx/.xlsm/.csv — the same
    universe tradebook_parser.py covers, since this is specifically a
    tradebook-shaped-data fallback, not a general document parser."""
    lower = filename.lower()
    rows: List[tuple] = []

    try:
        if lower.endswith((".xlsx", ".xlsm")):
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows.extend(list(ws.iter_rows(values_only=True)))
        elif lower.endswith((".csv", ".tsv")):
            import csv
            delimiter = "\t" if lower.endswith(".tsv") else ","
            text = None
            for enc in ("utf-8", "utf-8-sig", "latin-1"):
                try:
                    text = file_bytes.decode(enc)
                    break
                except UnicodeDecodeError:
                    continue
            if text is None:
                return {"transactions": [], "warnings": [], "error": "Could not decode this "
                        "file as text (unrecognized character encoding).", "columns_matched": []}
            rows = list(csv.reader(io.StringIO(text), delimiter=delimiter))
        else:
            return {"transactions": [], "warnings": [], "error": "This parser only handles "
                    ".xlsx, .xlsm, .csv, or .tsv files.", "columns_matched": []}
    except Exception as e:
        return {"transactions": [], "warnings": [], "error": f"Could not read this file: {e}",
                "columns_matched": []}

    header_result = _find_header_row(rows)
    if header_result is None:
        return {"transactions": [], "warnings": [], "error": "No confident tradebook-style "
                "header row found (need at least Symbol + Quantity + Price [or separate Buy "
                "Price/Sell Price columns], plus a Date or Buy/Sell column). This file's layout "
                "doesn't look like trade data, or uses column names too different from what "
                "this parser recognizes.", "columns_matched": []}
    header_row_idx = header_result["row_index"]
    col_map = header_result["columns"]
    split_price = "price" not in col_map  # i.e. relying on buy_price/sell_price instead

    def get(row, field):
        idx = col_map.get(field)
        return row[idx] if idx is not None and idx < len(row) else None

    transactions = []
    skipped = 0
    for row in rows[header_row_idx + 1:]:
        if row is None or get(row, "symbol") is None:
            continue
        symbol = str(get(row, "symbol")).strip()
        if not symbol or _normalize_header(symbol) in {"total", "grandtotal", "subtotal"}:
            continue

        qty = _to_float(get(row, "quantity"))
        trade_type = _parse_trade_type(get(row, "trade_type")) if "trade_type" in col_map else None
        trade_date = _parse_date(get(row, "date")) if "date" in col_map else None

        if split_price:
            # Some brokers (confirmed against a real export) use separate Buy
            # Price / Sell Price columns, populating only the one that
            # applies to that row's direction. Pick based on trade_type;
            # if trade_type itself is unknown, try whichever column has a
            # value, preferring buy_price.
            buy_p = _to_float(get(row, "buy_price"))
            sell_p = _to_float(get(row, "sell_price"))
            if trade_type == "buy":
                price = buy_p
            elif trade_type == "sell":
                price = sell_p
            else:
                price = buy_p if buy_p is not None else sell_p
        else:
            price = _to_float(get(row, "price"))

        if qty is None or price is None or trade_type is None or trade_date is None:
            skipped += 1
            continue

        isin = str(get(row, "isin")).strip() if get(row, "isin") else None
        if isin and not _ISIN_RE.match(isin):
            isin = None

        transactions.append({
            "symbol": symbol, "isin": isin, "trade_date": trade_date,
            "exchange": None, "segment": None, "trade_type": trade_type,
            "quantity": qty, "price": price, "trade_id": None,
            "order_id": None, "executed_at": None,
        })

    warnings = []
    if not transactions:
        warnings.append("A tradebook-style header was found, but no rows below it had all of "
                         "Symbol/Quantity/Price/Date/Type filled in — nothing was extracted.")
    elif skipped:
        warnings.append(f"{skipped} row(s) below the header were skipped — missing quantity, "
                         "price, a recognizable date, or a recognizable buy/sell value.")

    return {"transactions": transactions, "warnings": warnings, "error": None,
            "columns_matched": sorted(col_map.keys())}


def new_batch_id() -> str:
    return uuid.uuid4().hex[:12]
