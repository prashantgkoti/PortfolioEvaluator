"""
tradebook_parser.py — Parses Zerodha Console tradebook exports (.xlsx) into
individual buy/sell transactions, and computes FIFO cost basis per ISIN.

This fills a real gap: NSDL/CDSL CAS statements carry current value but
never purchase cost for demat equities (see cas_parser.py's DECISIONS #4).
A tradebook is the actual trade-by-trade record, so FIFO cost basis
computed from it is exact for whatever portion of the trading history has
been uploaded — not an estimate.

DECISION: built and verified against Zerodha Console's tradebook template
specifically (a real 411-row, 68-symbol file) — a "Client ID" label, a
"Tradebook for <Segment> from <date> to <date>" title, and a fixed column
set (Symbol, ISIN, Trade Date, Exchange, Segment, Series, Trade Type,
Auction, Quantity, Price, Trade ID, Order ID, Order Execution Time). The
header row is located by scanning for a row containing both "Symbol" and
"ISIN", not a hardcoded row number, so minor template variations (extra
blank rows, a different title) don't break it. Motilal Oswal and Angel One
tradebook exports use different column layouts and aren't covered by this
parser yet — a known limitation, same spirit as cas_parser.py's.

DECISION: Zerodha caps a single tradebook export at 365 days, so investors
with multi-year history must upload several files over time. FIFO
positions are therefore always computed over ALL transactions currently
stored in the database (across every upload), not just the most recent
file — otherwise a position opened in an earlier upload would look
phantom-empty. Duplicate trades (identified by Zerodha's own Trade ID) are
skipped on re-upload rather than double-counted.
"""
from __future__ import annotations

import uuid
from collections import deque
from typing import List

import openpyxl

from . import asset_classifier


def parse_tradebook_bytes(file_bytes: bytes) -> dict:
    """Returns {"transactions": [...], "warnings": [...], "error": str|None,
    "client_id": str|None, "title": str|None}."""
    try:
        import io
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        return {"transactions": [], "warnings": [], "error": f"Could not open this file as an "
                f"Excel workbook: {e}", "client_id": None, "title": None}

    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    client_id, title, header_row_idx, header_map = None, None, None, {}
    for i, row in enumerate(rows):
        for j, cell in enumerate(row):
            if cell == "Client ID" and j + 1 < len(row):
                client_id = row[j + 1]
            if isinstance(cell, str) and cell.startswith("Tradebook for"):
                title = cell
        if row and any(c == "Symbol" for c in row) and any(c == "ISIN" for c in row):
            header_row_idx = i
            header_map = {j: c for j, c in enumerate(row) if c is not None}
            break

    if header_row_idx is None:
        return {"transactions": [], "warnings": [], "error": "This doesn't look like a Zerodha "
                "Console tradebook export — couldn't find a header row with 'Symbol' and 'ISIN' "
                "columns.", "client_id": client_id, "title": title}

    name_to_col = {v: k for k, v in header_map.items()}
    required = ["Symbol", "ISIN", "Trade Date", "Trade Type", "Quantity", "Price"]
    missing = [r for r in required if r not in name_to_col]
    if missing:
        return {"transactions": [], "warnings": [], "error": f"Missing expected column(s): "
                f"{', '.join(missing)}. This tradebook format may differ from what's supported.",
                "client_id": client_id, "title": title}

    def get(row, col_name, default=None):
        idx = name_to_col.get(col_name)
        return row[idx] if idx is not None and idx < len(row) else default

    transactions = []
    for row in rows[header_row_idx + 1:]:
        if not row or get(row, "Symbol") is None:
            continue
        try:
            symbol = str(get(row, "Symbol")).strip()
            isin = str(get(row, "ISIN")).strip() if get(row, "ISIN") else None
            transactions.append({
                "symbol": symbol,
                "isin": isin,
                "asset_type": asset_classifier.classify(isin, symbol),
                "trade_date": str(get(row, "Trade Date")),
                "exchange": get(row, "Exchange"),
                "segment": get(row, "Segment"),
                "trade_type": str(get(row, "Trade Type", "")).strip().lower(),
                "quantity": float(get(row, "Quantity", 0) or 0),
                "price": float(get(row, "Price", 0) or 0),
                "trade_id": str(get(row, "Trade ID")) if get(row, "Trade ID") else None,
                "order_id": str(get(row, "Order ID")) if get(row, "Order ID") else None,
                "executed_at": str(get(row, "Order Execution Time")) if get(row, "Order Execution Time") else None,
            })
        except (ValueError, TypeError):
            continue

    warnings = []
    if not transactions:
        warnings.append("No transaction rows were found below the header. The file may be empty "
                         "or use a layout this parser doesn't recognize.")

    return {"transactions": transactions, "warnings": warnings, "error": None,
            "client_id": client_id, "title": title}


def compute_fifo_positions(transactions: List[dict]) -> dict:
    """Walks all transactions in date order and returns current net positions
    per instrument: {key: {symbol, isin, asset_type, quantity, avg_cost,
    realized_pnl}}. Grouped by ISIN when available, falling back to the
    symbol string when it isn't — some real broker exports (confirmed: an
    Angel One-style trade history report) have no ISIN column at all, and
    grouping strictly by ISIN would silently drop every one of those
    transactions from analysis entirely, contradicting the basic
    requirement that every recorded transaction — including a stock bought
    and fully sold — actually gets analyzed, not just stored. Fully exited
    positions (net quantity == 0) are still included, with avg_cost=None,
    so realized P&L on closed positions remains visible."""
    by_key: dict = {}
    for t in transactions:
        key = t.get("isin") or t.get("symbol")
        if key:
            by_key.setdefault(key, []).append(t)

    positions = {}
    for key, trades in by_key.items():
        trades_sorted = sorted(trades, key=lambda t: (t["trade_date"], t.get("executed_at") or ""))
        lots = deque()  # [quantity_remaining, price] per open buy lot, oldest first
        realized_pnl = 0.0
        symbol = trades_sorted[0]["symbol"]
        isin = trades_sorted[0].get("isin")
        asset_type = trades_sorted[0].get("asset_type") or "stock"

        # DECISION: each transaction gets a "matched_quantity" — for buys,
        # always the full quantity (every buy is real capital going in,
        # regardless of what's later sold). For sells, only the portion
        # actually matched against a known open lot. A sell whose quantity
        # exceeds what's actually available in the lot queue (confirmed on
        # real data: a position with several same-day sell rows totalling
        # more than any known prior buy) means shares predate the uploaded
        # trade history — that unmatched excess is excluded from
        # matched_quantity so XIRR cash-flow construction doesn't treat a
        # phantom, never-purchased sell as a real cash inflow.
        cashflow_txns = []
        for t in trades_sorted:
            qty, price = t["quantity"], t["price"]
            if t["trade_type"] == "buy":
                lots.append([qty, price])
                cashflow_txns.append({**t, "matched_quantity": qty})
            elif t["trade_type"] == "sell":
                remaining = qty
                matched_total = 0.0
                while remaining > 1e-9 and lots:
                    lot_qty, lot_price = lots[0]
                    matched = min(lot_qty, remaining)
                    realized_pnl += matched * (price - lot_price)
                    lots[0][0] -= matched
                    remaining -= matched
                    matched_total += matched
                    if lots[0][0] <= 1e-9:
                        lots.popleft()
                # If `remaining` is still >0 here, sells exceed known buys in the
                # uploaded history (e.g. shares held before this tradebook's date
                # range, or transferred in from another broker) — that excess is
                # simply not matched against any known cost, and ignored rather
                # than driving the position negative.
                cashflow_txns.append({**t, "matched_quantity": matched_total})

        total_qty = sum(l[0] for l in lots)
        if total_qty > 1e-6:
            total_cost = sum(l[0] * l[1] for l in lots)
            positions[key] = {"symbol": symbol, "isin": isin, "asset_type": asset_type,
                               "quantity": round(total_qty, 4), "avg_cost": round(total_cost / total_qty, 4),
                               "realized_pnl": round(realized_pnl, 2), "transactions": cashflow_txns}
        elif abs(realized_pnl) > 1e-6:
            positions[key] = {"symbol": symbol, "isin": isin, "asset_type": asset_type,
                               "quantity": 0.0, "avg_cost": None,
                               "realized_pnl": round(realized_pnl, 2), "transactions": cashflow_txns}

    return positions


def new_batch_id() -> str:
    return uuid.uuid4().hex[:12]
