"""
llm_parser.py — Fallback extraction for transaction/holdings reports that
don't match any of the deterministic parsers (cas_parser.py's NSDL layouts,
tradebook_parser.py's Zerodha Console format). Instead of hardcoding a new
parser per broker/format, this sends the document's extracted text to an
LLM and asks it to return structured holdings or transactions matching a
fixed schema — so a Motilal Oswal statement, an Angel One ledger, a CAMS/
KFintech mutual fund statement, or any other report can be ingested without
writing a bespoke parser for each one.

============================================================================
IMPORTANT — READ BEFORE ENABLING
============================================================================
This is the ONE place in the app that sends financial document content
outside the local machine. Every other parser in this codebase (cas_parser,
tradebook_parser) runs entirely locally — nothing leaves the machine.
This module calls the Anthropic API with the extracted text of whatever
file it's given, which may contain account numbers, holdings, and personal
identifying information present in the source document.

Because of that, this path is:
  - OFF by default (see db.AppSettings.llm_parsing_enabled) — the person
    running the app must explicitly enable it.
  - Used only as a FALLBACK, after the deterministic parsers have already
    failed to recognize the file — known formats never take this path.
  - Dependent on an ANTHROPIC_API_KEY environment variable being set on the
    machine running the backend. It is never stored in the database, never
    accepted from the frontend, and never logged.
============================================================================

DECISION: extraction uses Claude's tool-use (forced tool_choice) rather than
asking for JSON in prose, since a forced tool call is validated against a
schema by the API itself and is far more reliable to parse than hoping a
free-text response contains well-formed JSON.

DECISION: documents are chunked at ~12,000 characters with the LLM invoked
once per chunk, and results merged with best-effort deduplication —
holdings by (name, isin), transactions by (symbol, trade_date, trade_type,
quantity, price) — since a single very large statement could otherwise
exceed a reasonable single-call context and cost.

DECISION: the model never receives instructions to compute or infer values
it isn't given — the prompt explicitly tells it to use null for anything
not clearly present in the text, rather than estimating, to avoid
fabricated numbers entering the portfolio silently.

Untested against a live API call in this build — the extraction pipeline
(text extraction, chunking, prompt/schema construction, response
validation, and merge/dedup logic) is exercised with a mocked LLM response,
but the actual network round-trip to Anthropic's API has not been verified
end-to-end and should be checked with a real API key before relying on it.
"""
from __future__ import annotations

import io
import json
import os
import re
import uuid
from typing import List, Optional

DEFAULT_MODEL = os.environ.get("ANTHROPIC_MODEL", "claude-sonnet-5")
CHUNK_CHARS = 12_000
MAX_CHUNKS = 15  # hard cap so one enormous file can't trigger runaway API spend


def is_available() -> bool:
    """True only if the anthropic package is installed AND an API key is
    configured. Does not check the opt-in setting — callers check that
    separately (db.get_settings()['llm_parsing_enabled'])."""
    if not os.environ.get("ANTHROPIC_API_KEY"):
        return False
    try:
        import anthropic  # noqa: F401
        return True
    except ImportError:
        return False


# --------------------------------------------------------------------------- #
# Generic text extraction — any file type, not just PDF/XLSX
# --------------------------------------------------------------------------- #

def extract_text(filename: str, file_bytes: bytes) -> Optional[str]:
    lower = filename.lower()
    try:
        if lower.endswith(".pdf"):
            import pdfplumber
            with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                return "\n".join(page.extract_text() or "" for page in pdf.pages)

        if lower.endswith((".xlsx", ".xlsm")):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            lines = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                lines.append(f"=== Sheet: {sheet_name} ===")
                for row in ws.iter_rows(values_only=True):
                    if any(c is not None for c in row):
                        lines.append(" | ".join("" if c is None else str(c) for c in row))
            return "\n".join(lines)

        if lower.endswith((".csv", ".tsv", ".txt")):
            for enc in ("utf-8", "utf-8-sig", "latin-1"):
                try:
                    return file_bytes.decode(enc)
                except UnicodeDecodeError:
                    continue
            return file_bytes.decode("utf-8", errors="replace")

    except Exception:
        return None

    return None  # unrecognized extension for text extraction


def _chunk_text(text: str) -> List[str]:
    if len(text) <= CHUNK_CHARS:
        return [text]
    chunks = []
    start = 0
    while start < len(text) and len(chunks) < MAX_CHUNKS:
        end = min(start + CHUNK_CHARS, len(text))
        # try to break on a line boundary so a row doesn't get split mid-line
        if end < len(text):
            nl = text.rfind("\n", start, end)
            if nl > start:
                end = nl
        chunks.append(text[start:end])
        start = end
    return chunks


# --------------------------------------------------------------------------- #
# LLM extraction
# --------------------------------------------------------------------------- #

EXTRACTION_TOOL = {
    "name": "record_financial_data",
    "description": "Records structured holdings and/or transactions found in a financial document excerpt.",
    "input_schema": {
        "type": "object",
        "properties": {
            "document_type": {
                "type": "string",
                "enum": ["holdings", "transactions", "mixed", "unknown"],
                "description": "'holdings' = a point-in-time snapshot of positions (quantity, current value). "
                                "'transactions' = a trade-by-trade record (individual buy/sell events with dates). "
                                "'mixed' = both appear in this excerpt. 'unknown' = this excerpt is not a "
                                "financial holdings/transaction record (e.g. it's a cover letter, T&Cs, or blank).",
            },
            "holdings": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "name": {"type": "string"},
                        "symbol": {"type": ["string", "null"]},
                        "isin": {"type": ["string", "null"]},
                        "asset_type": {"type": "string", "enum": ["stock", "mutual_fund", "etf", "other"]},
                        "quantity": {"type": ["number", "null"]},
                        "avg_cost": {"type": ["number", "null"], "description": "Purchase cost per unit, only if explicitly stated — never estimated."},
                        "current_price": {"type": ["number", "null"]},
                        "current_value": {"type": ["number", "null"]},
                        "currency": {"type": "string", "default": "INR"},
                    },
                    "required": ["name", "asset_type"],
                },
            },
            "transactions": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "symbol": {"type": "string"},
                        "isin": {"type": ["string", "null"]},
                        "trade_date": {"type": "string", "description": "ISO format YYYY-MM-DD"},
                        "trade_type": {"type": "string", "enum": ["buy", "sell"]},
                        "quantity": {"type": "number"},
                        "price": {"type": "number"},
                    },
                    "required": ["symbol", "trade_date", "trade_type", "quantity", "price"],
                },
            },
            "notes": {"type": "string", "description": "Brief note on parsing confidence, ambiguity, or anything skipped."},
        },
        "required": ["document_type", "holdings", "transactions"],
    },
}

SYSTEM_PROMPT = """You extract structured holdings and transactions from financial statement text (bank/broker/depository exports, in any format or language — Indian markets terminology expected: NSE/BSE, ISIN, demat, folio, NAV, etc.).

Rules:
- Only extract data actually present in the text. Never estimate, infer, or compute a value that isn't stated.
- Use null for any field not clearly present, rather than guessing.
- If a row's meaning is ambiguous, skip it rather than recording something you're not confident about — note the skip in `notes`.
- A "holding" is a snapshot position (has quantity and current value/price, no per-trade date). A "transaction" is an individual trade event (has a specific trade date and buy/sell direction).
- Distinguish mutual funds (folio numbers, NAV, AMC names) from direct equities (NSE/BSE symbols, ISIN starting with the issuer's equity code) where possible.
- Always call the record_financial_data tool with your findings, even if both arrays end up empty."""


def _call_llm(text_chunk: str, filename: str) -> dict:
    """Returns {"document_type", "holdings", "transactions", "notes", "error"}."""
    import anthropic

    client = anthropic.Anthropic()  # reads ANTHROPIC_API_KEY from env
    try:
        response = client.messages.create(
            model=DEFAULT_MODEL,
            max_tokens=4096,
            system=SYSTEM_PROMPT,
            tools=[EXTRACTION_TOOL],
            tool_choice={"type": "tool", "name": "record_financial_data"},
            messages=[{
                "role": "user",
                "content": f"Filename: {filename}\n\nDocument excerpt:\n\n{text_chunk}",
            }],
        )
    except anthropic.AuthenticationError:
        return {"error": "ANTHROPIC_API_KEY is set but was rejected by the API — check it's valid."}
    except anthropic.RateLimitError:
        return {"error": "Anthropic API rate limit hit — try again shortly."}
    except anthropic.APIConnectionError:
        return {"error": "Could not reach the Anthropic API — check network connectivity."}
    except Exception as e:
        return {"error": f"LLM extraction failed: {e}"}

    for block in response.content:
        if block.type == "tool_use" and block.name == "record_financial_data":
            data = block.input
            return {
                "document_type": data.get("document_type", "unknown"),
                "holdings": data.get("holdings", []) or [],
                "transactions": data.get("transactions", []) or [],
                "notes": data.get("notes", ""),
                "error": None,
            }

    return {"error": "The model didn't return structured data as expected."}


_ISIN_RE = re.compile(r"^IN[A-Z0-9]{10}$")


def _normalize_holding(h: dict) -> Optional[dict]:
    if not h.get("name"):
        return None
    isin = h.get("isin")
    if isin and not _ISIN_RE.match(isin):
        isin = None  # drop anything that isn't a plausible ISIN rather than trust it blindly
    return {
        "symbol": h.get("symbol"), "name": h["name"], "isin": isin,
        "asset_type": h.get("asset_type") if h.get("asset_type") in ("stock", "mutual_fund", "etf", "other") else "other",
        "market": "IN", "quantity": h.get("quantity"), "avg_cost": h.get("avg_cost"),
        "current_price": h.get("current_price"), "current_value": h.get("current_value"),
        "currency": h.get("currency") or "INR", "notes": "AI-extracted — verify against the source document.",
    }


def _normalize_transaction(t: dict) -> Optional[dict]:
    if not t.get("symbol") or not t.get("trade_date") or t.get("trade_type") not in ("buy", "sell"):
        return None
    if t.get("quantity") is None or t.get("price") is None:
        return None
    isin = t.get("isin")
    if isin and not _ISIN_RE.match(isin):
        isin = None
    return {
        "symbol": t["symbol"], "isin": isin, "trade_date": str(t["trade_date"]),
        "exchange": None, "segment": None, "trade_type": t["trade_type"],
        "quantity": float(t["quantity"]), "price": float(t["price"]),
        "trade_id": None, "order_id": None, "executed_at": None,
    }


def extract_financial_data(filename: str, file_bytes: bytes) -> dict:
    """Top-level entry point. Returns {"holdings": [...], "transactions": [...],
    "document_type": str, "notes": [...], "error": str|None, "chunks_processed": int}."""
    if not is_available():
        return {"holdings": [], "transactions": [], "document_type": "unknown", "notes": [],
                "error": "AI-assisted parsing isn't available: set the ANTHROPIC_API_KEY "
                         "environment variable on the machine running the backend, and make "
                         "sure the 'anthropic' package is installed.", "chunks_processed": 0}

    text = extract_text(filename, file_bytes)
    if not text or not text.strip():
        return {"holdings": [], "transactions": [], "document_type": "unknown", "notes": [],
                "error": "No extractable text found in this file (it may be a scanned image, "
                         "an unsupported binary format, or empty).", "chunks_processed": 0}

    chunks = _chunk_text(text)
    all_holdings, all_transactions, all_notes = [], [], []
    doc_types = []

    for chunk in chunks:
        result = _call_llm(chunk, filename)
        if result.get("error"):
            return {"holdings": all_holdings, "transactions": all_transactions,
                    "document_type": "unknown", "notes": all_notes,
                    "error": result["error"], "chunks_processed": len(all_notes)}
        doc_types.append(result["document_type"])
        for h in result["holdings"]:
            normalized = _normalize_holding(h)
            if normalized:
                all_holdings.append(normalized)
        for t in result["transactions"]:
            normalized = _normalize_transaction(t)
            if normalized:
                all_transactions.append(normalized)
        if result.get("notes"):
            all_notes.append(result["notes"])

    # dedupe holdings by (name, isin); transactions by full-row identity
    seen_h, deduped_h = set(), []
    for h in all_holdings:
        key = (h["name"], h["isin"])
        if key not in seen_h:
            seen_h.add(key)
            deduped_h.append(h)

    seen_t, deduped_t = set(), []
    for t in all_transactions:
        key = (t["symbol"], t["trade_date"], t["trade_type"], t["quantity"], t["price"])
        if key not in seen_t:
            seen_t.add(key)
            deduped_t.append(t)

    overall_type = "mixed" if deduped_h and deduped_t else (
        "holdings" if deduped_h else ("transactions" if deduped_t else "unknown"))

    return {"holdings": deduped_h, "transactions": deduped_t, "document_type": overall_type,
            "notes": all_notes, "error": None, "chunks_processed": len(chunks)}


def new_batch_id() -> str:
    return uuid.uuid4().hex[:12]
